import openpyxl as op
from openpyxl.styles import PatternFill

NUMS = []
while True :
    n = input("그룹의 수를 입력하세요 (입력을 마치려면 0을 입력하세요) >> ")
    if n == '0' :
        print("입력이 완료되었습니다. ")
        print("작업중... 잠시만 기다려주세요!")
        break
    NUMS.append(n)

# 데이터 불러오기
frontierFile = op.load_workbook("FrontierPeoples-List-checkdup.xlsx")
unreachedFile = op.load_workbook("UnreachedPeoples-List-checkdup.xlsx")
ws_frontier = frontierFile["all"]
ws_unreached = unreachedFile["all"]
all_frontier = ws_frontier["A2:AG4891"]
all_unreached = ws_unreached["A2:AG7281"]

# 새 엑셀 시트 생성 
new_exc = op.Workbook()
all = new_exc.create_sheet("all")
new_exc.remove(new_exc["Sheet"])

ws_UPF = new_exc.create_sheet("UPG")
ws_FPG = new_exc.create_sheet("FPG")

ws_India = new_exc.create_sheet("India")
ws_China = new_exc.create_sheet("China")

colName = ["ROG3","Ctry","PeopleID3","ROP3","PeopNameAcrossCountries","PeopNameInCountry","Population","JPScale","LeastReached","ROL3","PrimaryLanguageName","BibleStatus","RLG3","PrimaryReligion","PercentAdherents",'PercentEvangelical','PeopleID1','ROP1','AffinityBloc','PeopleID2','ROP2','PeopleCluster','CountOfCountries','RegionCode','RegionName','ROG2','Continent','10_40Window','IndigenousCode','WorkersNeeded','Frontier','Latitude',"longitude"]
extractedColName = ["순번","나라명","종족명","종교","인원수","미전도/프론티어",'비고']
# 프론티어 그룹 아이디 세트 생성
frontier_list = set()
for fr in all_frontier :
    frontier_list.add((fr[0].value,fr[2].value))
idx = [2,1,2,1,2,1,2,1]

for h in range(1,len(colName) + 1) :
    all.cell(1,h,colName[h-1])
    ws_India.cell(1,h,colName[h-1])
    ws_China.cell(1,h,colName[h-1])
    ws_FPG.cell(1,h,colName[h-1])
    ws_UPF.cell(1,h,colName[h-1])

for i in range(len(all_unreached)) :
    for j in range(len(all_unreached[i])):
        all.cell(i+2,j+1,all_unreached[i][j].value)

    if all_unreached[i][1].value == "India" :
        for j in range(len(all_unreached[i])):
            ws_India.cell(idx[0],idx[1],all_unreached[i][j].value)
            idx[1] += 1
        idx[0] += 1
        idx[1] = 1

    elif all_unreached[i][1].value == "China" :
        for j in range(len(all_unreached[i])):
            ws_China.cell(idx[2],idx[3],all_unreached[i][j].value)
            idx[3] += 1
        idx[2] += 1
        idx[3] = 1
    elif (all_unreached[i][0].value,all_unreached[i][2].value) in frontier_list :
        for j in range(len(all_unreached[i])):
            ws_FPG.cell(idx[4],idx[5],all_unreached[i][j].value)
            idx[5] += 1
        idx[4] +=1
        idx[5] = 1
    else :
        for j in range(len(all_unreached[i])):
            ws_UPF.cell(idx[6],idx[7],all_unreached[i][j].value)
            idx[7] += 1
        idx[6] +=1
        idx[7] = 1

print("작업중...")

for num in NUMS :
    ws_groups = new_exc.create_sheet(num + "_Groups")
    ws_extreacted = new_exc.create_sheet(num + "_Groups_요약")
    NUM = int(num)
    groups = [[] for _ in range(NUM)]
    india = ws_India["A2:AG" + str(idx[0]-1)]
    china = ws_China["A2:AG" + str(idx[2]-1)]
    fpg = ws_FPG["A2:AG" + str(idx[4]-1)]
    upf = ws_UPF["A2:AG" + str(idx[6]-1)]
    index = 0

    for grp in india :
        groups[index%NUM].append(grp)
        index += 1
    for grp in china :
        groups[index%NUM].append(grp)
        index += 1
    for grp in fpg :
        groups[index%NUM].append(grp)
        index += 1
    for grp in upf :
        groups[index%NUM].append(grp)
        index += 1

    ii = 1
    groupId = 1
    m = len(groups[0])
    count = 0
    for group in groups :
        ws_groups.cell(ii,1,"group " + str(groupId))
        ws_extreacted.cell(ii,1,"group " + str(groupId))
        ii += 1
        for h in range(1,len(colName) + 1) :
            ws_groups.cell(ii,h,colName[h-1])

        for h in range(1,len(extractedColName) + 1) :
            ws_extreacted.cell(ii,h,extractedColName[h-1])

        ii += 1
        count = 0
        for i in range(len(group)):
            count += 1
            if group[i][1].value == "India" :
                fill = fill = PatternFill(start_color="FFDAB9", end_color="FFDAB9", fill_type="solid")
            elif group[i][1].value == "China" :
                fill = fill = PatternFill(start_color="EEE8AA", end_color="EEE8AA", fill_type="solid")
            elif group[i][30].value == "Y" :
                fill = fill = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
            elif group[i][30].value == "N" :
                fill = fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
            
            for j in range(len(group[i])):
                cell = ws_groups.cell(ii,j+1,group[i][j].value)
                cell.fill = fill

            cell = ws_extreacted.cell(ii,1,i)
            cell.fill = fill
            ex_idx = [1,4,13,6,30]
            for j in range(len(ex_idx)) :
                if ex_idx[j] != 30 :
                    cell = ws_extreacted.cell(ii,j+2,group[i][ex_idx[j]].value)
                    cell.fill = fill
                else :
                    if group[i][ex_idx[j]].value == "Y" :
                        cell = ws_extreacted.cell(ii,j+2,"FPG")
                    else :
                        cell = ws_extreacted.cell(ii,j+2,"UPF")
                    cell.fill = fill
            ii += 1
        ii += 2
        groupId += 1
        if count < m :
            ii += 1
    print("작업중...")

new_exc.save("result.xlsx")
print("작업이 완료되었습니다!")