import openpyxl as op

NUM = 861

# 데이터 불러오기
frontierFile = op.load_workbook("FrontierPeoples-List-checkdup.xlsx")
unreachedFile = op.load_workbook("UnreachedPeoples-List-checkdup.xlsx")
ws_frontier = frontierFile["all"]
ws_unreached = unreachedFile["all"]
all_frontier = ws_frontier["A2:AG4891"]
all_unreached = ws_unreached["A2:AG7281"]

# 새 엑셀 시트 생성 
new_exc = op.Workbook()
all = new_exc.create_sheet("all")
new_exc.remove(new_exc["Sheet"])

ws_UPF = new_exc.create_sheet("UPG")
ws_FPG = new_exc.create_sheet("FPG")

ws_India = new_exc.create_sheet("India")
ws_China = new_exc.create_sheet("China")
new_exc[0].columns()
# 프론티어 그룹 아이디 세트 생성
frontier_list = set()
for fr in all_frontier :
    frontier_list.add((fr[0].value,fr[2].value))
idx = [1,1,1,1,1,1,1,1]


for i in range(len(all_unreached)) :
    for j in range(len(all_unreached[i])):
        all.cell(i+1,j+1,all_unreached[i][j].value)

    if all_unreached[i][1].value == "India" :
        for j in range(len(all_unreached[i])):
            ws_India.cell(idx[0],idx[1],all_unreached[i][j].value)
            idx[1] += 1
        idx[0] += 1
        idx[1] = 1

    elif all_unreached[i][1].value == "China" :
        for j in range(len(all_unreached[i])):
            ws_China.cell(idx[2],idx[3],all_unreached[i][j].value)
            idx[3] += 1
        idx[2] += 1
        idx[3] = 1
    elif (all_unreached[i][0].value,all_unreached[i][2].value) in frontier_list :
        for j in range(len(all_unreached[i])):
            ws_FPG.cell(idx[4],idx[5],all_unreached[i][j].value)
            idx[5] += 1
        idx[4] +=1
        idx[5] = 1
    else :
        for j in range(len(all_unreached[i])):
            ws_UPF.cell(idx[6],idx[7],all_unreached[i][j].value)
            idx[7] += 1
        idx[6] +=1
        idx[7] = 1

groups = [[] for _ in range(NUM)]

new_exc.save("result.xlsx")