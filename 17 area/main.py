import openpyxl as op
from openpyxl.styles import PatternFill

NUM = 5
filename = "area17"
# 데이터 불러오기
xl_area = op.load_workbook(filename + ".xlsx")
ws_area = xl_area[filename]

groups = [[] for _ in range(NUM)]
index = 0

for i in range(1,ws_area.max_row + 1) :
    groups[(i-1)%NUM].append(ws_area[i])

for gr in groups :
    print(len(gr))

print(ws_area.max_row)

ws_result = xl_area.create_sheet("result")

groupId = 1
m = len(groups[0])
count = 0
ii = 1
extractedColName = ["순번","나라명","종족명","종교","인원수","미전도/프론티어",'비고']
ex_idx = [1,5,10,13,12]
for group in groups :
    ws_result.cell(ii,1,"group " + str(groupId))
    ii += 1

    for h in range(1,len(extractedColName) + 1) :
        ws_result.cell(ii,h,extractedColName[h-1])

    ii += 1
    count = 0
    for i in range(1,len(group)):
        count += 1       
        cell = ws_result.cell(ii,1,i)
        
        for j in range(len(ex_idx)) :
            if ex_idx[j] != 30 :
                cell = ws_result.cell(ii,j+2,group[i][ex_idx[j]].value)
            else :
                if group[i][ex_idx[j]].value == "Y" :
                    cell = ws_result.cell(ii,j+2,"FPG")
                else :
                    cell = ws_result.cell(ii,j+2,"UPF")

        ii += 1
    ii += 2
    groupId += 1
    if count < m :
        ii += 1


xl_area.save(filename + ".xlsx")